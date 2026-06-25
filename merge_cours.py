#!/usr/bin/env python3
"""
merge_cours.py — met à jour le cache « Tous les cours » (cache_cours.xlsx) en
fusionnant l'historique existant avec un nouvel export AEC récent.

Pourquoi : l'export AEC « Cours > Tous les cours » qu'on télécharge ne couvre en
pratique que l'année scolaire courante (≈ sept. N → août N+1). Le cache de prod
(`oscar-prealpha/web/server/data/new_cours/cache_cours.xlsx`) est lui un CUMUL
pluriannuel (2022→…). Remplacer bêtement effacerait l'historique. Ce script :

- garde du cache actuel toutes les lignes dont « Date début » < cutoff
  (ou sans date) → historique intact ;
- prend du nouvel export toutes les lignes dont « Date début » >= cutoff
  → année courante complète (remplace une éventuelle version tronquée).

Le schéma « Tous les cours » (154 colonnes) est préservé positionnellement (les
en-têtes du cache, dédupliqués par pandas, font foi). Les dates sont réécrites
comme vraies dates Excel, lisibles par `aec_parser_v3`.

Usage :
    python merge_cours.py <cache_actuel.xlsx> <nouvel_export.xlsx> <sortie.xlsx> [YYYY-MM-DD]

Cutoff par défaut : 2025-09-01 (rentrée de l'année scolaire 25-26). Pensez à
adapter le cutoff à la rentrée de l'année que vous rafraîchissez. Après fusion,
remplacez cache_cours.xlsx puis (optionnel) régénérez le fixture :
    python oscar-prealpha/web/server/build_snapshot.py
"""
import sys
from datetime import datetime
from openpyxl import load_workbook, Workbook

SHEET = "AEC"


def pick_sheet(wb):
    return wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[0]]


def to_dt(v):
    if isinstance(v, datetime):
        return v
    if v is None:
        return None
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(2)
    cur_path, new_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    cutoff = datetime.strptime(sys.argv[4], "%Y-%m-%d") if len(sys.argv) > 4 else datetime(2025, 9, 1)

    cur_wb = load_workbook(cur_path, read_only=True, data_only=True)
    new_wb = load_workbook(new_path, read_only=True, data_only=True)
    cur_ws, new_ws = pick_sheet(cur_wb), pick_sheet(new_wb)

    cur_rows = list(cur_ws.iter_rows(values_only=True))
    new_rows = list(new_ws.iter_rows(values_only=True))
    cur_header, new_header = list(cur_rows[0]), list(new_rows[0])

    # Les colonnes doivent s'aligner POSITIONNELLEMENT. Le cache a des en-têtes
    # dédupliqués par pandas ('Nom du cours.1') et strippés ; l'export brut a les
    # doublons + espaces finaux. On tolère ces écarts cosmétiques connus.
    if len(cur_header) != len(new_header):
        print(f"ABORT: nb colonnes différent ({len(cur_header)} vs {len(new_header)})")
        sys.exit(1)

    def base(s):
        s = (str(s).strip() if s is not None else "")
        if "." in s and s.rsplit(".", 1)[-1].isdigit():
            s = s.rsplit(".", 1)[0]
        return s

    bad = [(i, a, b) for i, (a, b) in enumerate(zip(cur_header, new_header)) if base(a) != base(b)]
    if bad:
        print("ABORT: colonnes non alignées (au-delà des doublons/espaces connus) :")
        for i, a, b in bad[:20]:
            print(f"  col {i}: cache='{a}'  vs  new='{b}'")
        sys.exit(1)
    print(f"Alignement colonnes OK ({len(cur_header)} colonnes)")

    norm = [str(h).strip() if h is not None else "" for h in cur_header]
    date_idx = norm.index("Date début")

    kept_hist, dropped_recent = [], 0
    for row in cur_rows[1:]:
        dt = to_dt(row[date_idx])
        if dt is None or dt < cutoff:
            kept_hist.append(row)
        else:
            dropped_recent += 1

    taken_new, skipped_old = [], 0
    for row in new_rows[1:]:
        dt = to_dt(row[date_idx])
        if dt is not None and dt >= cutoff:
            taken_new.append(row)
        else:
            skipped_old += 1

    print(f"Cache actuel : {len(cur_rows) - 1} lignes -> historique gardé (<{cutoff:%Y-%m-%d}) = {len(kept_hist)}, retirées (>=cutoff) = {dropped_recent}")
    print(f"Nouvel export: {len(new_rows) - 1} lignes -> prises (>=cutoff) = {len(taken_new)}, ignorées (<cutoff) = {skipped_old}")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = SHEET
    out_ws.append(cur_header)
    for r in kept_hist:
        out_ws.append(list(r))
    for r in taken_new:
        out_ws.append(list(r))
    out_wb.save(out_path)
    print(f"OK -> {out_path}  ({len(kept_hist) + len(taken_new)} lignes de données, {len(cur_header)} colonnes)")


if __name__ == "__main__":
    main()
